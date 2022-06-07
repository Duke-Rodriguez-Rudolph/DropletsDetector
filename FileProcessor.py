import openpyxl
import os
import numpy as np

class FileProcessor:
    # 构造函数
    def __init__(self,excel_path):
        self.__path=excel_path    # 存储excel文件的位置
        #如果excel文件不存在则新建，如果存在则读取
        if os.path.exists(self.__path)==False:
            self.__wb=openpyxl.Workbook()
            self.__sheet = self.__wb.active
            self.__sheet.title='all'
            self.__wb.create_sheet('clusters')
            self.__wb.create_sheet('result')
            self.__sheet.append(['列_行','颜色','是否发光','x1','y1','w1','h1','x2','y2','w2','h2'])
            for i in range(0,200):
                for j in range(0,50):
                    n=i*50+j+2
                    self.__sheet.cell(n, 1).value=str(i)+'_'+str(j)
            self.__wb.save(self.__path)
        else:
            self.__wb=openpyxl.load_workbook(self.__path)
            self.__sheet = self.__wb['all']

    # 写入位置信息(行，列，位置信息几，信息)
    def __writeLocation(self,row,col,index,location):
        n = col * 50 + row + 1 + 1
        [x,y,w,h,area]=location
        box=[x,y,w,h]
        for i,key in enumerate(box):
            self.__sheet.cell(n, i+4+index*4).value = key

    # 写入颜色信息(行，列，颜色信息)
    def __writeColor(self,row,col,color):
        n = col * 50 + row + 1 + 1
        self.__sheet.cell(n, 2).value = color

    # 写入发光信息(行，列，是否发光)
    def __writeBlink(self,row,col,is_blink):
        n=col*50+row+1+1
        self.__sheet.cell(n, 3).value=is_blink

    # 写入荧光识别
    def writeBlink(self,blink_list,none_list):
        for blink in blink_list:
            col,row=blink.split('_')
            self.__writeBlink(int(row),int(col),1)
        for none in none_list:
            col,row=none.split('_')
            self.__writeBlink(int(row),int(col),0)
        self.__wb.save(self.__path)

    # 写入颜色分类
    def writeColor(self,droplets_dict):
        for droplet_type in droplets_dict:
            droplet_type_dict = droplets_dict[droplet_type]
            for jpg in droplet_type_dict:
                # 拆分
                col, row = jpg.split('_')
                self.__writeColor(int(row),int(col),droplet_type)
        self.__wb.save(self.__path)

    # 写入分割信息
    def writeLocation(self,droplets,index):
        for col, droplet in enumerate(droplets):
            for row,droplet_small in enumerate(droplet):
                self.__writeLocation(int(row),int(col),index,droplet_small)
        self.__wb.save(self.__path)

    # 写入聚类中心
    def writeClusters(self,clusters):
        self.__sheet = self.__wb['clusters']
        self.__sheet.append(['颜色', 'b', 'g','r'])
        for i,color in enumerate(clusters):
            b, g, r = color
            self.__sheet.append([i, b, g, r])
        self.__sheet = self.__wb['all']
        self.__wb.save(self.__path)

    # 读取信息
    def readInformation(self,is_color):
        droplets_dict={}
        blink_list=[]
        none_list=[]
        droplets1=[]
        droplet1=[]
        droplets2 = []
        droplet2 = []
        colors={}
        for row in self.__sheet.iter_rows(min_row=2):
            # 赋值
            droplet_type=row[1].value
            is_blink=row[2].value
            adroplet1 = [row[3].value, row[4].value, row[5].value, row[6].value,row[5].value*row[6].value]
            adroplet2 = [row[7].value, row[8].value, row[9].value, row[10].value,row[9].value*row[10].value]
            # 颜色分类
            droplet_type_dict=droplets_dict.get(droplet_type,[])
            droplet_type_dict.append(row[0].value)
            droplets_dict[droplet_type]=droplet_type_dict
            # 荧光检测
            if is_blink:
                blink_list.append(row[0].value)
            else:
                none_list.append(row[0].value)
            # 液滴归位
            droplet1.append(adroplet1)
            droplet2.append(adroplet2)
            if len(droplet1)==50:
                droplet1 = np.array(droplet1)
                droplet2 = np.array(droplet2)
                droplets1.append(droplet1)
                droplets2.append(droplet2)
                droplet1 = []
                droplet2 = []

        if is_color:
            self.__sheet = self.__wb['clusters']
            for row in self.__sheet.iter_rows(min_row=2):
                colors[row[0].value]=(row[1].value,row[2].value,row[3].value)
            self.__sheet = self.__wb['all']

        return droplets1,droplets2,droplets_dict,blink_list,none_list,colors

    # 制作结果
    def doResult(self):
        droplets1, droplets2, droplets_dict, blink_list, none_list, colors = self.readInformation(0)
        self.__sheet = self.__wb['result']
        self.__sheet.append(['颜色', '总数量', '阳性数量'])

        for droplet_type in droplets_dict:
            if droplet_type=='none':
                continue
            all_num=len(droplets_dict[droplet_type])
            print('类别'+str(droplet_type)+'的总数量为：',all_num)
            blink_num=0
            for jpg in droplets_dict[droplet_type]:
                if jpg in blink_list:
                    blink_num+=1
            print('类别' + str(droplet_type) + '的阳性数量为：', blink_num)
            self.__sheet.append([droplet_type, all_num, blink_num])
        self.__sheet.append(['none', len(droplets_dict['none']), '/'])
        print('类别none的总数量为：', len(droplets_dict['none']))

        self.__sheet = self.__wb['all']
        self.__wb.save(self.__path)
